from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import time
import openpyxl
from selenium.webdriver.common.keys import Keys

workbook = openpyxl.load_workbook("../Excels/Student_Example.xlsx")
sheet = workbook['Student']
rows = sheet.max_row
columns = sheet.max_column

for r in range(1,rows+1):
    for c in range(1, columns+1):
        print(sheet.cell(r,c).value)


driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))

driver.get("https://demoqa.com/automation-practice-form")
driver.maximize_window()
driver.implicitly_wait(5)

ciacho = driver.find_element(By.CLASS_NAME, "fc-button-label")
ciacho.click()

First_Name = driver.find_element(By.ID, 'firstName')
First_Name.send_keys(sheet.cell(1,1).value)

Second_Name = driver.find_element(By.ID, "lastName")
Second_Name.send_keys(sheet.cell(2,1).value)

Email_Addr = driver.find_element(By.ID, "userEmail")
Email_Addr.send_keys(sheet.cell(3,1).value)

#Gender choose

if sheet.cell(4,1).value == ("male" or "Male"):
    Male_Button = driver.find_element(By.XPATH, "//*[@id='genterWrapper']/div[2]/div[1]/label")
    Male_Button.click()
elif sheet.cell(4,1).value == ("female" or "Female"):
    Female_Button = driver.find_element(By.XPATH, "//*[@id='genterWrapper']/div[2]/div[2]/label")
    Female_Button.click()
else:
    Other_Button = driver.find_element(By.XPATH, "//*[@id='genterWrapper']/div[2]/div[3]/label")
    Other_Button.click()

Phone_Number = driver.find_element(By.ID,'userNumber')
Phone_Number.send_keys(sheet.cell(5,1).value)

#Date of Birth

# Callendar_pick= driver.find_element(By.ID, "dateOfBirthInput")
# Callendar_pick.click()
#
# select_year = driver.find_element(By.CLASS_NAME, "react-datepicker__year-select")
# select_year.send_keys(1999)
# select_year_month = driver.find_element(By.CLASS_NAME, "react-datepicker__month-select")
# select_year_month.send_keys(3)
# select_day = driver.find_element(By.CLASS_NAME, "react-datepicker__day react-datepicker__day--025 react-datepicker__day--selected")
# select_day.send_keys(25)

time.sleep(5)

#Hobbies choice
Sports_hobby = driver.find_element(By.ID, "hobbies-checkbox-1")
Reading_hobby = driver.find_element(By.ID, "hobbies-checkbox-2")
Music_hobby = driver.find_element(By.XPATH, "//*[@id='hobbiesWrapper']/div[2]/div[3]/label")


# hobbies_value = sheet.cell(8, 1).value
# if "Sports" in hobbies_value:
#     Sports_hobby.click()
# elif "Reading" in hobbies_value:
#     Reading_hobby.click()
# elif "Music" in hobbies_value:
#     Music_hobby.click()
# else:
#     print("No valid hobby chosen")

Picture_upl = driver.find_element(By.ID, 'uploadPicture')
Picture_upl.send_keys(sheet.cell(9,1).value)

Curr_Addr = driver.find_element(By.ID,"currentAddress")
Curr_Addr.send_keys(sheet.cell(10,1).value)

time.sleep(5)



# Select_State = driver.find_element(By.ID, "state")
# Select_State.click()

# Select_City = driver.find_element(By.ID, "city")
# Select_City.click()



Scroll_Submit = driver.find_element(By.ID,"submit")
Scroll_Submit.send_keys(Keys.END)
Scroll_Submit.click()

time.sleep(5)

print()
print("TEST PASS")
driver.quit()